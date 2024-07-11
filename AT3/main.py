import math
import argparse

import numpy as np
import pandas as pd
import scipy.stats as stats
import matplotlib.pyplot as plt

from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


def show_dict(data_dict):
    for k in data_dict:
        print(f"{k} {data_dict[k]}")

def sturges_formula(n):
    return 1 + 3.322 * math.log(n, 10)

def get_class_interval(minimun, number_of_class, c):
    return [(minimun + i * c, minimun + (i + 1) * c) for i in range(number_of_class)]

def calculate_xis(intervals):
    return [(a + b) / 2.0 for a, b in intervals]

def calculate_freq(data, intervals):
    freqs = [0 for _ in intervals]
    for d in data:
        for i, interval in enumerate(intervals):
            if interval[0] <= d < interval[1]:
                freqs[i] += 1
                break
    return freqs[:]

def calculate_pis(freqs, n):
    return [freq / n for freq in freqs]

def calculate_xi_dot_pi(xis, pis):
    return [xi * pi for xi, pi in zip(xis, pis)]

def calculate_qdp(xis, pis, mean):
    return [(xi - mean)**2 * pi for xi, pi in zip(xis, pis)]

def detect_outliers(data):
    Q1 = np.percentile(data, 25)
    Q3 = np.percentile(data, 75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    outliers = data[(data < lower_bound) | (data > upper_bound)]
    return outliers, lower_bound, upper_bound

def calculate_descriptive_model(data):
    n = len(data)
    sturges = sturges_formula(n)
    number_of_class = sturges
    minimum = data.min()
    maximum = data.max()
    r = maximum - minimum
    c =  round(r / number_of_class, 4)
    mean = data.mean()
    median = data.median()
    std_deviation_value = data.std()
    moda = data.mode()[0]
    class_intervals = get_class_interval(minimum, math.ceil(number_of_class), c)
    xis = calculate_xis(class_intervals)
    freqs = calculate_freq(data, class_intervals)
    pis = calculate_pis(freqs, n)
    Pis = [sum(pis[:i + 1]) for i in range(len(pis))]
    xis_dot_pis = calculate_xi_dot_pi(xis, pis)
    weighted_average = sum(xis_dot_pis)
    qdp = calculate_qdp(xis, pis, mean)
    variance_value = sum(qdp)
    outliers, lower_bound, upper_bound = detect_outliers(data)
    relative_error = abs(mean - weighted_average) / mean
    CV = std_deviation_value / weighted_average
    assi = (weighted_average - moda) / std_deviation_value
    dist = data.skew()

    return {
        "dados": list(data),
        "numero de dados": n,
        "numero de classes": math.ceil(number_of_class),
        "max": maximum,
        "min": minimum,
        "range": r,
        "c": c,
        "media": mean,
        "mediana": median,
        "var": variance_value,
        "desvpad": math.sqrt(sum(qdp)),
        "classes": class_intervals,
        "ni": freqs,
        "pi": pis,
        "Pi": Pis,
        "Xi": xis,
        "Xi*pi": xis_dot_pis,
        "QDP": qdp,
        "media ponderada": weighted_average,
        "erro relativo": relative_error,
        "CV": CV,
        "assimetria": assi,
        "distorcao": dist,
        "limite inferior": lower_bound,
        "limite superior": upper_bound,
        "outliers": outliers.tolist()
    }

def calculate_zs(intervals, mean, std):
    return zip(*[((a - mean) / std, (b - mean) / std) for a, b in intervals])

def calculate_probability(z1s, z2s):
    probabilitys = [stats.norm.cdf(z2) - stats.norm.cdf(z1) for z1, z2 in zip(z1s, z2s)]
    # fazendo as extremidades irem de menos infinito ao primeiro numero do intervalo
    # e do ultimo numero do intervalo a mais infinito
    probabilitys[0] += stats.norm.cdf(z1s[0])
    probabilitys[-1] = stats.norm.cdf(float("+inf")) - stats.norm.cdf(z1s[-1])
    return probabilitys

def calculate_es(p_z1s_z2s, n):
    return [n * p_z1_z2 for p_z1_z2 in p_z1s_z2s]

def calculate_qui_squares(freqs, es):
    return [(freq - e)**2 / e for freq, e in zip(freqs, es)]

def adjust(intervals, freqs, es, error=5.0):
    adjust_intervals = []
    adjust_freqs = []
    adjust_es = []

    i = 0
    acc_intervals = intervals[0]
    acc_freqs = freqs[i]
    acc_es = es[i]
    i += 1
    while True:
        if acc_es >= error:
            break
        acc_es += es[i]
        acc_freqs += freqs[i]
        acc_intervals = (acc_intervals[0], intervals[i][1])
        i += 1


    adjust_intervals.append(acc_intervals)
    adjust_freqs.append(acc_freqs)
    adjust_es.append(acc_es)

    j = len(intervals) - 1
    acc_intervals = intervals[-1]
    acc_freqs = freqs[j]
    acc_es = es[j]
    j -= 1
    while True:
        if acc_es >= error:
            break
        acc_es += es[j]
        acc_freqs += freqs[j]
        acc_intervals = (intervals[j][0], acc_intervals[1])
        j -= 1

    for h in range(i, j + 1):
        adjust_intervals.append(intervals[h])
        adjust_freqs.append(freqs[h])
        adjust_es.append(es[h])

    adjust_intervals.append(acc_intervals)
    adjust_freqs.append(acc_freqs)
    adjust_es.append(acc_es)

    return adjust_intervals, adjust_freqs, adjust_es

def chi_square(n, class_intervals, mean, std_deviation_value, freqs):

    z1s, z2s = calculate_zs(class_intervals, mean, std_deviation_value)
    p_z1s_z2s = calculate_probability(z1s, z2s)
    es = calculate_es(p_z1s_z2s, n)
    adjust_intervals, adjust_freqs, adjust_es = adjust(class_intervals, freqs, es)
    qui_squares = calculate_qui_squares(adjust_freqs, adjust_es)
    df = len(adjust_intervals) - 2 - 1
    chi2_critical = stats.chi2.ppf(0.95, df)

    return {
        "classes": class_intervals,
        "Oi": freqs,
        "z1": z1s,
        "z2": z2s,
        "P(z1<Z<z2)": p_z1s_z2s,
        "Ei": es,
        "classes ajustadas": adjust_intervals,
        "Oi ajustado": adjust_freqs,
        "Ei ajustada": adjust_es,
        "QQI": qui_squares,
        "sum(Oi)": sum(freqs),
        "sum(P(z1<Z<z2))": sum(p_z1s_z2s),
        "sum(Ei)": sum(es),
        "sum(Oi ajustado)": sum(adjust_freqs),
        "sum(Ei ajustada)": sum(adjust_es),
        "Qui-quadrado": sum(qui_squares),
        "Qui-quadrado critico": chi2_critical,
    }

def dict2xlsx(data_dict, name="output.xlsx"):
    # Preparação do DataFrame
    # Primeiro, transformar valores únicos em listas do mesmo tamanho
    max_length = max(len(v) if isinstance(v, list) else 1 for v in data_dict.values())
    for key, value in data_dict.items():
        if not isinstance(value, list) and not isinstance(value, tuple):
            data_dict[key] = [value] + [""] * (max_length-1)
        else:
            # Se a lista for menor que o comprimento máximo, completar com None
            data_dict[key] = list(data_dict[key]) + [""] * (max_length - len(data_dict[key]))

    # Conversão do dicionário para DataFrame
    df = pd.DataFrame(data_dict)

    # Arredondando
    def round_value(item, decimals=4):
        # Se o item for uma tupla, arredondar cada elemento numérico dentro da tupla
        if isinstance(item, tuple):
            return tuple(round(x, decimals) if isinstance(x, (int, float)) else x for x in item)
        # Se o item for um número (int ou float), arredondar diretamente
        elif isinstance(item, (int, float)):
            return round(item, decimals)
        # Se o item não for nem tupla nem número, retornar o item sem modificação
        return item

    # Função para formatar a tupla no formato 'x[0] |- x[1]'
    def format_interval(interval):
        if isinstance(interval, tuple) or isinstance(interval, list):
            return f"{interval[0]:.4f} |- {interval[1]:.4f}"
        return interval

    # Aplicando a função em uma coluna que contém tuplas com possíveis valores numéricos
    for col in df.columns:
        df[col] = df[col].apply(round_value)    

    if 'classes' in df.columns:
        # Aplicando a função à coluna 'A' se ela existir
        df['classes'] = df['classes'].apply(format_interval)

    if 'classes ajustadas' in df.columns:
        # Aplicando a função à coluna 'A' se ela existir
        df['classes ajustadas'] = df['classes ajustadas'].apply(format_interval)

    # Criação de um Workbook do Openpyxl e adição de uma Sheet
    wb = Workbook()
    ws = wb.active

    # Preenchimento da Sheet com dados do DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=str(value))
            if r_idx == 1:  # Aplicar a formatação na primeira linha
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Ajustar largura da coluna
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in col))
        adjusted_width = (max_length + 2) * 1.1  # Ajuste para melhor visualização
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Alinhando
    for cell in ws["1:1"]:  # "1:1" indica que a formatação será aplicada à primeira linha
        cell.alignment = Alignment(horizontal="center")

    # Definir o estilo de borda em negrito
    borda_negrito = Side(border_style="thin", color="000000")  # "thick" para borda em negrito
    borda_completa = Border(top=borda_negrito, left=borda_negrito, right=borda_negrito, bottom=borda_negrito)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = borda_completa

    # Salvar o Workbook
    wb.save(name)


def graphs(data, class_number, path=None):
    # Criação da figura e subplots
    fig, axs = plt.subplots(1, 3, figsize=(18, 5))

    # Histograma
    axs[0].hist(data, bins=class_number, alpha=0.7, rwidth=0.85)
    axs[0].set_title('Histograma dos Dados')

    # Q-Q Plot
    stats.probplot(data, dist="norm", plot=axs[1])
    axs[1].set_title('Q-Q Plot dos Dados')

    # Boxplot
    axs[2].boxplot(data, vert=True)
    axs[2].set_title('Boxplot dos Dados')
    axs[2].set_xticks([])  # Remove os ticks do eixo y

    # Ajuste dos layouts
    plt.tight_layout()

    if path:
        plt.savefig(path)

    plt.show()

def part_1(sample1, sample2, name1, name2):
    print("PARTE 1")

    model1 = calculate_descriptive_model(sample1)
    print(f"Modelo descritivo {name1}")
    print()
    show_dict(model1)
    print()

    model2 = calculate_descriptive_model(sample2)
    print(f"Modelo descritivo {name2}")
    print()
    show_dict(model2)
    print()
    print("FIM DA PARTE 1")
    print()

    dict2xlsx(model1, f"datas/modelo-descritivo-{name1}.xlsx")
    dict2xlsx(model2, f"datas/modelo-descritivo-{name2}.xlsx")

    return model1, model2

def part_2(sample1, sample2, name1, name2):
    def partial_analysis(sample, name):
        model = calculate_descriptive_model(sample)
        n = model["numero de dados"]
        number_of_class = model["numero de classes"]
        class_intervals = model["classes"]
        freqs = model["ni"]
        weighted_average = model["media ponderada"]
        std_deviation_value = model["desvpad"]

        chi_square_result = chi_square(n, class_intervals, weighted_average, std_deviation_value, freqs)
        dict2xlsx(chi_square_result.copy(), f"datas/teste-aderencia-{name}.xlsx")
        graphs(sample, number_of_class, f"datas/{name}.png")
        return chi_square_result

    result_1  = partial_analysis(sample1, name1 + "_normal")
    sample1 = np.log(sample1)
    result_lognormal_1 = partial_analysis(sample1, name1 + "_lognormal")

    print("PARTE 2")

    print(f"Teste de aderência normal {name1}")
    print()
    show_dict(result_1)
    print()

    print(f"Teste de aderência lognormal {name1}")
    print()
    show_dict(result_lognormal_1)
    print()

    result_2 = partial_analysis(sample2, name2 + "_normal")
    sample1 = np.log(sample2)
    result_lognormal_2 = partial_analysis(sample2, name2 + "_lognormal")

    print(f"Teste de aderência normal {name2}")
    print()
    show_dict(result_2)
    print()

    print(f"Teste de aderência lognormal {name2}")
    print()
    show_dict(result_lognormal_2)
    print()

    print("FIM DA PARTE 2")
    print()

    return result_1, result_lognormal_1, result_2, result_lognormal_2

def variance_test(n1, n2, var1, var2, alpha=0.05):
    # Teste para a razão entre variâncias
    f_statistic = var1 / var2
    dfn, dfd = n1 - 1, n2 - 1
    f_critical1 = stats.f.ppf(alpha / 2, dfn, dfd)
    f_critical2 = stats.f.ppf(1 - alpha / 2, dfn, dfd)
    reject_var_h0 = f_statistic < f_critical1 or f_statistic > f_critical2

    # Construção do intervalo de confiança para a razão entre variâncias
    ci_var_lower = f_statistic / f_critical2
    ci_var_upper = f_statistic * f_critical2

    return {
        'Estatística F': f_statistic,
        'Valor crítico F (lower)': f_critical1,
        'Valor crítico F (upper)': f_critical2,
        'IC razão variâncias (lower)': ci_var_lower,
        'IC razão variâncias (upper)': ci_var_upper,
        'Rejeição H0 variâncias iguais': reject_var_h0,
    }

def average_test(n1, n2, mean1, mean2, var1, var2, alpha=0.05):
    # Teste para a diferença de médias (assumindo variâncias iguais)
    pooled_var = ((n1 - 1) * var1 + (n2 - 1) * var2) / (n1 + n2 - 2)
    t_statistic = (mean1 - mean2) / np.sqrt(pooled_var * (1 / n1 + 1 / n2))
    df = n1 + n2 - 2
    t_critical = stats.t.ppf(1 - alpha / 2, df)
    reject_mean_h0 = abs(t_statistic) > t_critical

    # Construção do intervalo de confiança para a diferença de médias
    ci_mean_lower = (mean1 - mean2) - t_critical * np.sqrt(pooled_var * (1 / n1 + 1 / n2))
    ci_mean_upper = (mean1 - mean2) + t_critical * np.sqrt(pooled_var * (1 / n1 + 1 / n2))

    return {
        's²': pooled_var,
        'Estatística t': t_statistic,
        'Valor crítico t': t_critical,
        'IC diferença médias (lower)': ci_mean_lower,
        'IC diferença médias (upper)': ci_mean_upper,
        'Rejeição H0 médias iguais': reject_mean_h0,
    }

def part_3(sample1, sample2):
    model1 = calculate_descriptive_model(sample1)
    model2 = calculate_descriptive_model(sample2)
    n1, n2 = model1["numero de dados"], model2["numero de dados"]
    mean1, mean2 = model1["media ponderada"], model2["media ponderada"]
    var1, var2 = model1["var"], model2["var"]

    variance_test_result = variance_test(n1, n2, var1, var2)
    average_test_result = average_test(n1, n2, mean1, mean2, var1, var2)

    results = variance_test_result | average_test_result

    print()
    print("PARTE 3")
    print()

    for k in results:
        print(f"{k}: {results[k]}")

    print()
    print("FIM DA PARTE 3")

    dict2xlsx(results, "datas/estatisticas-parametricas.xlsx")

    return results

def main():

    parser = argparse.ArgumentParser(description="Lê um arquivo Excel e imprime na tela.")
    parser.add_argument("--excel_path", help="Caminho do arquivo Excel para ser lido.", default="./datas/BASE DADOS_DESAFIO INDIVIDUAL.xlsx")
    args = parser.parse_args()

    def filter_excel(df, filters):
        all_filter = pd.Series([True] * len(df), index=df.index)

        for f in filters:
            all_filter &= f(df)

        return df[all_filter]

    excel_path = args.excel_path
    dataframe = pd.read_excel(excel_path)

    embalagem = "C03 - VIDRO 600ML RET"
    produto = "CERVEJA"
    marca1 = "AMSTEL LAGER"
    marca2 = "ANTARCTICA PILSEN"
    seg1 = 2
    seg2 = 2

    filter_1 = [lambda df: df['Marca'] == marca1,
                lambda df: df[embalagem].notna(),
                lambda df: df["Seg"] == seg1,
                lambda df: df["Produto"] == produto]

    filter_2 = [lambda df: df['Marca'] == marca2,
                lambda df: df[embalagem].notna(),
                lambda df: df["Seg"] == seg2,
                lambda df: df["Produto"] == produto]

    sample1 = filter_excel(dataframe, filter_1)[embalagem]
    sample2 = filter_excel(dataframe, filter_2)[embalagem]

    part_1(sample1.copy(deep=True), sample2.copy(deep=True), marca1, marca2)
    part_2(sample1.copy(deep=True), sample2.copy(deep=True), marca1, marca2)
    part_3(sample1.copy(deep=True), sample2.copy(deep=True))


if __name__ == "__main__":
    main()
